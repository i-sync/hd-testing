"""
The market matrix excel utils
Get some matrix data from excel file. "MY20-RYI-Matrix-V3.1.xlsx"
"""

__author__ = "Michael.Tian"

import openpyxl

__matrix_file_name__ = "proj12_ryi/data/MY20-RYI-Matrix-V4.xlsx"

def get_social_matrix():
    """
    Get all soical link matrix
    locale column B
    Facebook Url column W
    Instagram Url column X
    Twitter Url column Y
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["RYI"]
    res = {}
    for index in range(4, 35):
        locale = sheet["B{}".format(index)].value.split()[0]
        facebook = sheet["W{}".format(index)].value
        instagram = sheet["X{}".format(index)].value
        twitter = sheet["Y{}".format(index)].value
        res[locale] = [link for link in [facebook, instagram, twitter] if link]

    wb.close()
    return res

def get_all_locale():
    """
    Get all locale from matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["RYI"]

    res = []
    for index in range(4, 35):
        locale = sheet["B{}".format(index)].value.split()[0]
        res.append(locale)

    return res

def get_bike_matrix():
    """
    Get All bike info matrix
    Custom: K,L,M,N,O
    Performance: P,Q,R,S
    Touring: T,U,V
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["RYI"]
    bike_category = {
        0: ["K", "L", "M", "N", "O"],
        1: ["P", "Q", "R", "S"],
        2: ["T", "U", "V"]
    }

    # bike_column = ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]
    bike_column = [col for c in [bike_category[key] for key in bike_category.keys()] for col in c]
    bike_code = {}
    for col in bike_column:
        code = sheet["{}1".format(col)].value
        bike_code[col] = code

    res = {}
    for index in range(4, 35):
        locale = sheet["B{}".format(index)].value.split()[0]
        res[locale] = {}
        for key in bike_category.keys():
            res[locale][key] = []
            column = bike_category[key]
            for col in column:
                cc = sheet["{}{}".format(col, index)].value
                if cc and not cc.startswith("N"):
                    res[locale][key].append(bike_code[col])

    wb.close()
    return res


if __name__ == "__main__":
    print(sorted(get_all_locale()))