"""
The market matrix excel utils
Get some matrix data from excel file. "MY20-Demo-Matrix-V6.5.xlsx"
"""

__author__ = "Michael.Tian"

import openpyxl

__matrix_file_name__ = "proj13_demo/data/MY20-Demo-Matrix-V6.5.xlsx"

def get_social_matrix():
    """
    Get all soical link matrix
    locale column C
    Facebook Url column AG
    Instagram Url column AH
    Twitter Url column AI
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["DEMO"]
    res = {}
    for index in range(4, 45):
        locale = sheet["C{}".format(index)].value.split()[0]
        facebook = sheet["AG{}".format(index)].value
        instagram = sheet["AH{}".format(index)].value
        twitter = sheet["AI{}".format(index)].value
        res[locale] = [link for link in [facebook, instagram, twitter] if link]

    wb.close()
    return res

def get_all_locale():
    """
    Get all locale from matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["DEMO"]

    res = []
    for index in range(4, 45):
        locale = sheet["C{}".format(index)].value.split()[0]
        res.append(locale)

    return res

def get_bike_matrix():
    """
    Get All bike info matrix
    HD Street: L,M
    Sportster: N,O,P,Q
    Softail: R,S,T,U,V,W,X
    Touring: Y,Z,AA,AB,AC,AD,AE
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["DEMO"]
    bike_category = {
        0: ["L", "M"],
        1: ["N", "O", "P", "Q"],
        2: ["R", "S", "T", "U", "V", "W", "X"],
        3: ["Y", "Z", "AA", "AB", "AC", "AD", "AE"]
    }

    # bike_column = ["K", "L", "M", "N", "O", "P", "Q", "R", "S", "T", "U", "V"]
    bike_column = [col for c in [bike_category[key] for key in bike_category.keys()] for col in c]
    bike_code = {}
    for col in bike_column:
        code = sheet["{}2".format(col)].value
        bike_code[col] = code

    res = {}
    for index in range(4, 45):
        locale = sheet["C{}".format(index)].value.split()[0]
        res[locale] = {}
        i = 0
        for key in bike_category.keys():
            #res[locale][key] = []
            tmp = []
            column = bike_category[key]
            for col in column:
                cc = sheet["{}{}".format(col, index)].value
                if cc and cc.startswith("Y"):
                    #res[locale][key].append(bike_code[col])
                    tmp.append(bike_code[col])
            if len(tmp):
                res[locale][i] = tmp
                i = i + 1

    wb.close()
    return res