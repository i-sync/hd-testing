"""
The market matrix excel utils
Get some footer links data from excel file. "MY19R_Social_footer_link.xlsx"
Get bike data from excel file. "MY19_FXDR_Market_Matrix.xlsx"
"""

__author__ = "Libby.Qin"

import openpyxl
__matrix_file_name__ = "proj04_touringdemo/data/MY19R_Social_footer_link.xlsx"
__matrix_bike_file_name__ = "proj04_touringdemo/data/MY19_FXDR_Market_Matrix.xlsx"
__matrix_share_copy__ = "proj04_touringdemo/data/MY19R_menu_share_copy.xlsx"

def get_social_footer_matrix():
    """
    Get all soical link matrix
    locale column B
    Facebook Url column C
    Instagram Url column D
    Twitter Url column E
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_file_name__)
    sheet = wb["MY19R"]
    res = {}
    for index in range(2, 44):
        locale = sheet["B{}".format(index)].value
        facebook = sheet["C{}".format(index)].value
        instagram = sheet["D{}".format(index)].value
        twitter = sheet["E{}".format(index)].value
        res[locale] = [facebook, instagram, twitter]
    wb.close()
    return res
def get_bike_matrix():
    """
    Get All bike info matrix
    :return:
    """
    wb = openpyxl.load_workbook(__matrix_bike_file_name__)

    sheet = wb["MY19R"]
    bike_column =["H", "I", "J", "K", "L", "M", "N", "O"]
    bike_code = {}
    for col in bike_column:
        id = sheet["{}1".format(col)].value
        id=id.upper()
        bike_code[col]= id

    res = {}
    for index in range(2, 34):
        locale = sheet["B{}".format(index)].value
        res[locale] = []
        for col in bike_column:
            cc = sheet["{}{}".format(col, index)].value
            if cc and  not cc.startswith("NO"):
                res[locale].append(bike_code[col])

    wb.close()
    return res

def Marge(d1,d2):
    ress={**d1,**d2}
    return ress
def get_menu_links_matrix():
    """
       Get All menu links matrix
       :return:
    """
    wb = openpyxl.load_workbook(__matrix_bike_file_name__)
    cp = openpyxl.load_workbook(__matrix_share_copy__)
    sheet = wb["MY19R"]
    sheet_cp=cp["MY19R"]
    res = {}
    share = {}
    for index in range(2, 33):
        locale = sheet["B{}".format(index)].value
        home = sheet["E{}".format(index)].value
        booking = sheet["F{}".format(index)].value
        res[locale] = [home, booking]

    for index in range(2,33):
        locale_cp = sheet_cp["B{}".format(index)].value
        copy_message = sheet_cp["C{}".format(index)].value
        # twitter = "https://twitter.com/intent/tweet?hashtags=&text=+" + copy_message + "+&tw_p=tweetbutton&url=https://ridefree.harley-davidson.com/" + locale_cp
        facebook = "https://www.facebook.com/sharer/sharer.php?u=https://ridefree.harley-davidson.com/" + locale_cp + "&appid=500689006785520"
        share[locale_cp]=[facebook]
    #根据相同key，合并两个字典
    for i, j in share.items():
        if i in res.keys():
            res[i] += j
        else:
            res.update({f'{i}': share[i]})
    wb.close()
    cp.close()
    return res